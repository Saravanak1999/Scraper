from openpyxl import Workbook
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from tld import get_tld
#create/open the file
try:
    workbook = load_workbook(filename="sample.xlsx")
except:
    workbook = Workbook()
sheet = workbook.active
try: 
    from googlesearch import search 
except ImportError:  
    print("No module named 'google' found")
#insert Title and Link at the top row
sheet.cell(row=1, column=1).value= "Title" 
sheet.cell(row=1, column=2).value= "Link"
#count the already existing rows
a=0;
for row in sheet.rows:
    a=a+1;
#insert keywords
queries = ["Medtechinnovator", "startup competition", "startup award"]

i=2;
for query in queries:
  #check the 10 results of each query
  for j in search(query, tld="co.in", num=10, stop=10, pause=2):
    #check if the link already exists in the sheet
    k=0;
    for row in sheet.rows:
        if j==row[1].value:
            k=1
            break
        #removes multiple results from the same website
        try:
            info=get_tld(j,as_object=True)
            info1=get_tld(row[1].value,as_object=True)
            if info.parsed_url[1]==info1.parsed_url[1]:
                k=1
                break
        except:
            continue
    #if it is not already present on the sheet, find the title using BeautifulSoup
    if k==0:
        url=j 
        try:
            r=requests.get(url)
            html_content = r.text
            soup = BeautifulSoup(html_content, 'lxml')
            t=soup.title.string
            if(t=="403 Forbidden" or t=="Not Acceptable!"):
                headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}
                r = requests.get(url,headers=headers)
                html_content = r.text
                soup = BeautifulSoup(html_content, 'lxml')
                t=soup.title.string
        #if the function gives any error, Apeend the keyword with Unknown and Insert it on the excel sheet
        except:
            t=query+" (Unknown)"
        #insert the new elements at the end of the sheet
        sheet.cell(row=(a+i), column=1).value=t
        sheet.cell(row=(a+i), column=2).value=j
        i=i+1;
  a=a+2;
            
workbook.save(filename="sample.xlsx")
